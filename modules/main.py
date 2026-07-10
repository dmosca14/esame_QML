import os
import pandas as pd
from datetime import datetime
import zoneinfo 
from sklearn.metrics import f1_score, accuracy_score, classification_report

from modules.dataset_preparation import prepare_dataset
from modules.training import train_model
from modules.diagnostics import evaluate_model
from modules.metrics import calculate_kernel_target_alignment, calculate_geometric_coefficient

# ------------------------------------------------------------------------------------------------

# MAIN CODE THAT BUILDS THE QSVM.

def qsvm(dataset, config, kernel_modules):

    # DEFINING GLOBAL VARIABLES.

    global_results = []
    K_classic_train_reference = None

    # UNPACKING CONFIGURATION FILE.

    number_samples = config["number_samples"]
    number_features = config["number_features"]
    fraction_train = config["fraction_train"]
    max_qubits = config["max_qubits"]
    quantum_device = config["quantum_device"]

    # LOADING THE DATA.

    dataset_name = dataset["dataset_name"]
    dataset_original_data = dataset["dataset_original_data"]

    # PREPARING THE DATASET.

    train_set, test_set = prepare_dataset(dataset_original_data, number_samples, number_features, fraction_train)

    X_train, y_train = train_set
    X_test, y_test = test_set

    # TIME AND FOLDER SETUP.

    local_time_zone = zoneinfo.ZoneInfo("Europe/Rome")
    timestamp_now = datetime.now(local_time_zone)

    primary_folder_str = "run_results"
    day_str = timestamp_now.strftime(f"%Y%m%d")
    hour_str = timestamp_now.strftime(f"%H%M%S")
    day_and_dataset_str = f"{day_str}_{dataset_name}"
    day_and_dataset_and_hour_str = f"{day_str}_{dataset_name}_{hour_str}"

    run_folder_name = os.path.join(primary_folder_str, day_str, day_and_dataset_str, day_and_dataset_and_hour_str)
    os.makedirs(run_folder_name, exist_ok = True) 

    # TRAINING.

    for kernel_module in kernel_modules:

        # UNPACKING THE KERNEL NAME AND SPECIFIC FOLDER GENERATION.

        kernel_name = kernel_module.kernel_name

        day_and_dataset_and_hour_and_encoding_str = f"{day_str}_{dataset_name}_{hour_str}_{kernel_name}"
        
        folder_name = os.path.join(run_folder_name, day_and_dataset_and_hour_and_encoding_str) 
        os.makedirs(folder_name, exist_ok = True) 

        # TRAINING.

        (best_model, 
         best_model_kernel_hyperparameters, 
         best_model_SVM_hyperparameters, 
         best_model_K_train,
         best_model_K_test) = train_model(dataset_name = dataset_name,
                                          X_train = X_train,
                                          y_train = y_train,
                                          X_test = X_test,
                                          number_features = number_features,
                                          max_qubits = max_qubits,
                                          quantum_device = quantum_device,
                                          kernel_module = kernel_module,
                                          kernel_name = kernel_name,
                                          folder_name = folder_name)
        
        # CALCULATING THE PREDICTIONS ON THE TEST SET.

        y_pred = best_model.predict(best_model_K_test)

        accuracy = accuracy_score(y_test, y_pred)
        f1 = f1_score(y_test, y_pred, average = 'weighted')

        # EVALUATING THE BEST MODEL.

        evaluate_model(dataset_name = dataset_name,
                       y_train = y_train,
                       y_test = y_test,
                       y_pred = y_pred,
                       kernel_name = kernel_name,
                       folder_name = folder_name, 
                       best_model_K_train = best_model_K_train)

        # CALCULATING THE METRICS.

        # Calculating the KTA.

        current_KTA = calculate_kernel_target_alignment(best_model_K_train, y_train)

        # Calculating the geometric coefficient.

        if kernel_name == "classic_RBF_kernel":

            K_classic_train_reference = best_model_K_train
            geometric_coefficient = "N/A"

        else:
    
            geometric_coefficient = f"{calculate_geometric_coefficient(K_classic_train_reference, best_model_K_train):.4f}"

        # EXTRACTING DETAILED METRICS FOR THE EXCEL.

        report_dict = classification_report(y_test, y_pred, output_dict = True, zero_division = 0)

        # UPDATING GLOBAL RESULTS.

        global_results.append({
            "dataset_name": dataset_name,
            "kernel_name": kernel_name,
            "kernel_hyperparameters": str(best_model_kernel_hyperparameters),
            "SVM_hyperparameters": str(best_model_SVM_hyperparameters),
            "KTA": f"{current_KTA:.4f}",
            "geometric_coefficient": geometric_coefficient,
            "accuracy": f"{accuracy:.4f}",
            "F1 score": f"{f1:.4f}",
            "macro_precision": f"{report_dict['macro avg']['precision']:.4f}",
            "macro_recall": f"{report_dict['macro avg']['recall']:.4f}",
            "class_0_precision": f"{report_dict.get('0', {}).get('precision', 0.0):.4f}",
            "class_0_recall": f"{report_dict.get('0', {}).get('recall', 0.0):.4f}",
            "class_0_f1_score": f"{report_dict.get('0', {}).get('f1-score', 0.0):.4f}",
            "class_0_support": int(report_dict.get('0', {}).get('support', 0)),
            "class_1_precision": f"{report_dict.get('1', {}).get('precision', 0.0):.4f}",
            "class_1_recall": f"{report_dict.get('1', {}).get('recall', 0.0):.4f}",
            "class_1_f1_score": f"{report_dict.get('1', {}).get('f1-score', 0.0):.4f}",
            "class_1_support": int(report_dict.get('1', {}).get('support', 0))
        })

        # SAVING THE PARTIAL RESULTS OF THIS KERNEL IN EXCEL (code made by Gemini).

        partial_excel_path = os.path.join(run_folder_name, f"partial_results_{dataset_name}.xlsx")
        df_partial = pd.DataFrame(global_results)
        
        with pd.ExcelWriter(partial_excel_path, engine='openpyxl') as writer:
            df_partial.to_excel(writer, index=False, sheet_name='Results')
            sheet = writer.sheets['Results']
            
            from openpyxl.utils import get_column_letter
    
            for column_index, column_name in enumerate(df_partial.columns):
                max_length = max(df_partial[column_name].astype(str).map(len).max(), len(column_name)) + 2
                column_letter = get_column_letter(column_index + 1)
                sheet.column_dimensions[column_letter].width = max_length

    df_final = pd.DataFrame(global_results)
    print(f"\nFINAL RESULTS FOR {dataset_name.upper()}:\n")
    print(df_final.to_string(index=False))

    # SAVING THE FINAL SUMMARY (code made by Gemini).

    final_excel_path = os.path.join(run_folder_name, f"final_summary_{dataset_name}.xlsx")
    
    with pd.ExcelWriter(final_excel_path, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Results')
        sheet = writer.sheets['Results']
        
        from openpyxl.utils import get_column_letter 
        from openpyxl.styles import PatternFill

        for column_index, column_name in enumerate(df_final.columns):
            max_length = max(df_final[column_name].astype(str).map(len).max(), len(column_name)) + 2
            column_letter = get_column_letter(column_index + 1)
            sheet.column_dimensions[column_letter].width = max_length

        best_index = df_final["F1 score"].astype(float).idxmax()
        excel_row = best_index + 2 
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, len(df_final.columns) + 1):
            sheet.cell(row=excel_row, column=col).fill = yellow_fill