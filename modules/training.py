import os
import time
import numpy as np
import pandas as pd
from sklearn.svm import SVC
from sklearn.metrics import f1_score
from sklearn.model_selection import ParameterGrid, StratifiedKFold
from tqdm import tqdm
from modules.dataset_preparation import apply_PCA

# ------------------------------------------------------------------------------------------------

# MODEL TRAINING FUNCTION.

def train_model(dataset_name,
                X_train,
                y_train,
                X_test,
                number_features,
                max_qubits,
                quantum_device,
                kernel_module,
                kernel_name,
                folder_name):
    
    # START.

    print(f"TRAINING {kernel_name.replace('_', ' ').upper()} FOR \"{dataset_name.upper()}\" DATASET\n")

    # CONFIGURING THE CROSS VALIDATION.

    CV_fold_number = 5

    class_count = np.unique(y_train, return_counts = True)[1]
    effective_fold_number = min(CV_fold_number, int(class_count.min()))

    CV_validator = StratifiedKFold(n_splits = effective_fold_number, shuffle = True, random_state = 42)
    fold_indices = list(CV_validator.split(X_train, y_train)) 

    # EXTRACTING THE HYPERPARAMETERS OF THE KERNEL.

    # Usually these hyperparameters are "quantum", because they refer to the kernel.
    # The only exception to this is for the classical RBF kernel.

    kernel_hyperparameters = kernel_module.get_kernel_hyperparameters(number_features, max_qubits)
    kernel_hyperparameters_grid = list(ParameterGrid(kernel_hyperparameters))

    # DEFINING THE SVM HYPERPARAMETERS.

    SVM_hyperparameters = {"C": [1.0, 10.0, 100.0, 1000.0]}
    SVM_hyperparameters_grid = list(ParameterGrid(SVM_hyperparameters))

    # DEFINING THE GLOBAL VARIABLES FOR THE CICLE.

    best_model, best_score, best_std_score = None, -1.0, None
    best_model_kernel_hyperparameters, best_model_SVM_hyperparameters = {}, None
    training_results_history = [] 

    # CONFIGURING THE UPDATE BAR.

    total_combinations = len(kernel_hyperparameters_grid) * len(SVM_hyperparameters_grid)
    with tqdm(total = total_combinations, desc = f"Training", bar_format = "{l_bar}{bar}| {n_fmt}/{total_fmt} (hyperparameters combinations)") as barra:
        
        # START OF THE CICLE ON ALL THE KERNEL HYPERPARAMETERS POSSIBLE COMBINATIONS. 

        for kernel_hyperparameters_combination in kernel_hyperparameters_grid:

            kernel_start_time = time.time()
        
            kernel_folds = []

            # CALCULATING THE KERNEL (Cross Validation).
            
            for train_fold_indices, val_fold_indices in fold_indices:

                X_train_fold = X_train[train_fold_indices]
                y_train_fold = y_train[train_fold_indices]

                X_val_fold = X_train[val_fold_indices]
                y_val_fold = y_train[val_fold_indices]

                X_train_fold_PCA, X_val_fold_PCA = apply_PCA(X_train_fold, X_val_fold, number_features)
                
                K_train_fold = kernel_module.calculate_kernel(X_train_fold_PCA, 
                                                              X_train_fold_PCA,
                                                              number_features, 
                                                              quantum_device, 
                                                              **kernel_hyperparameters_combination)
                
                K_val_fold = kernel_module.calculate_kernel(X_train_fold_PCA, 
                                                            X_val_fold_PCA, 
                                                            number_features, 
                                                            quantum_device, 
                                                            **kernel_hyperparameters_combination).T

                kernel_folds.append((K_train_fold, K_val_fold, y_train_fold, y_val_fold))
            
            kernel_end_time = time.time()
            kernel_total_time = kernel_end_time - kernel_start_time

            # TRAINING THE CLASSICAL SVM WITH THE PRECOMPUTED KERNEL.

            for SVM_hyperparameters_combination in SVM_hyperparameters_grid:

                fold_scores = []
                
                for K_train_fold, K_val_fold, y_train_f, y_val_f in kernel_folds:

                    fold_model = SVC(kernel = "precomputed", **SVM_hyperparameters_combination)
                    fold_model.fit(K_train_fold, y_train_f)
                    val_fold_predictions = fold_model.predict(K_val_fold)
                    fold_scores.append(f1_score(y_val_f, val_fold_predictions, average = "macro"))

                average_val_score = float(np.mean(fold_scores)) 
                std_val_score = float(np.std(fold_scores))    

                # UPDATAING THE RESULT.

                current_result = {"kernel_name": kernel_name,
                                  "kernel_hyperparameters": kernel_hyperparameters_combination,
                                  "SVM_hyperparameters": SVM_hyperparameters_combination,
                                  "F1_score_macro(%)": round(average_val_score * 100, 2), 
                                  "F1_score_std_CV(%)": round(std_val_score * 100, 2), 
                                  "kernel_calculation_time(s)": round(kernel_total_time),
                                  }
                
                training_results_history.append(current_result)

                # DECIDING IF THIS MODEL IS BETTER THEN THE PREVIOUS ONE.
                
                if average_val_score > best_score:
                    best_score = average_val_score
                    best_std_score = std_val_score
                    best_model_kernel_hyperparameters = kernel_hyperparameters_combination
                    best_model_SVM_hyperparameters = SVM_hyperparameters_combination

                barra.update(1)

    # ONCE THE BEST MODEL HAS BEEN FOUND, WE TRAIN IT ON THE ENTIRE SET.
    
    X_train_PCA, X_test_PCA = apply_PCA(X_train, X_test, number_features)

    best_model_K_train = kernel_module.calculate_kernel(X_train_PCA,
                                                        X_train_PCA,
                                                        number_features,
                                                        quantum_device,
                                                        **best_model_kernel_hyperparameters)
    
    best_model_K_test = kernel_module.calculate_kernel(X_train_PCA,
                                                       X_test_PCA,
                                                       number_features,
                                                       quantum_device,
                                                       **best_model_kernel_hyperparameters).T

    best_model = SVC(kernel = "precomputed", **best_model_SVM_hyperparameters)
    best_model.fit(best_model_K_train, y_train)

    # SAVING THE RESULTS ON A EXCEL FILE (code made by Gemini).

    df = pd.DataFrame(training_results_history)

    df["kernel_hyperparameters"] = df["kernel_hyperparameters"].astype(str)
    df["SVM_hyperparameters"] = df["SVM_hyperparameters"].astype(str)

    excel_file_name = os.path.join(folder_name, f"training_history_{dataset_name}.xlsx")

    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        sheet = writer.sheets['Results']

        from openpyxl.utils import get_column_letter 
        from openpyxl.styles import PatternFill

        for column_index, column_name in enumerate(df.columns):
            max_length = max(
                df[column_name].astype(str).map(len).max(),
                len(column_name)
            ) + 2
            column_letter = get_column_letter(column_index + 1)
            sheet.column_dimensions[column_letter].width = max_length

        best_index = df["F1_score_macro(%)"].idxmax()
        excel_row = best_index + 2 
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, len(df.columns) + 1):
            sheet.cell(row=excel_row, column=col).fill = yellow_fill

    # RESULTS.
    
    return (best_model,
            best_model_kernel_hyperparameters, 
            best_model_SVM_hyperparameters,
            best_model_K_train, 
            best_model_K_test)