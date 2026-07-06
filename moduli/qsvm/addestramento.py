import os
import time
import numpy as np
import pandas as pd
from sklearn.svm import SVC
from sklearn.metrics import f1_score
from sklearn.model_selection import ParameterGrid, StratifiedKFold
from tqdm import tqdm
from ..preprocessing_dataset.preparazione_dataset import prepara_dataset

def addestramento(nome_encoding, train_set_raw, test_set_raw, numero_features, modulo_encoding, iperparametri_quantistici, iperparametri_classici, nome_cartella, nome_dataset, numero_fold_cv = 5):

    X_cv_raw, y_cv_raw = train_set_raw

    conteggio_classi = np.unique(y_cv_raw, return_counts = True)[1]
    numero_fold_effettivo = min(numero_fold_cv, int(conteggio_classi.min()))

    # Divisione in 5 fold.

    validatore_cv = StratifiedKFold(n_splits = numero_fold_effettivo, shuffle = True, random_state = 42)
    indici_fold = list(validatore_cv.split(X_cv_raw, y_cv_raw)) 

    griglia_quantistica = list(ParameterGrid(iperparametri_quantistici))
    griglia_classica = list(ParameterGrid(iperparametri_classici))

    miglior_modello, miglior_score, miglior_std_score = None, -1.0, None
    migliori_matrici_gram, miglior_set_adattato = None, None  
    migliori_iperparametri_quantistici, migliori_iperparametri_classici = {}, {} 
    storico_risultati_addestramento = [] 

    totale_combinazioni = len(griglia_quantistica) * len(griglia_classica)
    with tqdm(total = totale_combinazioni, desc = f"Addestramento ({numero_fold_effettivo}-fold CV)", bar_format = "{l_bar}{bar}| {n_fmt}/{total_fmt}\n") as barra:
        
        for combinazione_iperparametri_quantistici in griglia_quantistica:
            tempo_inizio_kernel = time.time()

            # Iterazione sui vari fold.
        
            matrici_gram_folds = []
            for indici_train_fold, indici_val_fold in indici_fold:

                # Estrazione dei fold raw.
                
                train_fold_raw = (X_cv_raw[indici_train_fold], y_cv_raw[indici_train_fold])
                val_fold_raw = (X_cv_raw[indici_val_fold], y_cv_raw[indici_val_fold])

                # Standardizzazione e applicazione della PCA.

                train_fold_pca, val_fold_pca, _ = prepara_dataset(train_fold_raw, val_fold_raw, numero_features)

                # Applicazione dell'encoding.

                _, matrici_gram_fold, _ = modulo_encoding.kernel(train_fold_pca, val_fold_pca, numero_features, **combinazione_iperparametri_quantistici)
                
                # matrici_gram_fold[0] è Train vs Train, [1] è Val vs Train.

                matrici_gram_folds.append((matrici_gram_fold[0], matrici_gram_fold[1], y_cv_raw[indici_train_fold], y_cv_raw[indici_val_fold]))
            
            tempo_kernel = time.time() - tempo_inizio_kernel 

            for combinazione_iperparametri_classici in griglia_classica:
                tempo_inizio_SVM = time.time()

                punteggi_fold = []

                # Utilizziamo i kernel già calcolati fold-by-fold.
                
                for K_train_fold, K_val_fold, y_train_f, y_val_f in matrici_gram_folds:
                    modello_fold = SVC(kernel = "precomputed", **combinazione_iperparametri_classici)
                    modello_fold.fit(K_train_fold, y_train_f)
                    predizioni_val_fold = modello_fold.predict(K_val_fold)
                    punteggi_fold.append(f1_score(y_val_f, predizioni_val_fold, average = "macro"))

                score_val_medio = float(np.mean(punteggi_fold)) # media sui 5 fold
                score_val_std = float(np.std(punteggi_fold))    # dev std sui 5 fold

                tempo_SVM = time.time() - tempo_inizio_SVM 
                tempo_totale_combinazione = tempo_kernel + tempo_SVM 

                risultato_corrente = {
                    "encoding": nome_encoding,
                    "iperparametri_quantistici": combinazione_iperparametri_quantistici,
                    "iperparametri_classici": combinazione_iperparametri_classici,
                    "F1_score_macro(%)": round(score_val_medio * 100, 2), 
                    "F1_score_std_cv(%)": round(score_val_std * 100, 2), 
                    "tempo_calcolo_kernel(s)": round(tempo_kernel),
                    "tempo_addestramento_validazione_SVM(s)": round(tempo_SVM),
                    "tempo_esecuzione(s)": round(tempo_totale_combinazione)
                }
                storico_risultati_addestramento.append(risultato_corrente)
                
                if score_val_medio > miglior_score:
                    miglior_score = score_val_medio
                    miglior_std_score = score_val_std
                    migliori_iperparametri_quantistici = combinazione_iperparametri_quantistici 
                    migliori_iperparametri_classici = combinazione_iperparametri_classici

                barra.update(1)

    # Riaddestramento finale del miglior modello sull'intero set.
    # Scaling e calcolo della PCA eseguito un'ultima volta, per l'intero set (Train + Test test finale)

    train_pca_finale, test_pca_finale, _ = prepara_dataset(train_set_raw, test_set_raw, numero_features)
    
    miglior_set_adattato, migliori_matrici_gram, _ = modulo_encoding.kernel(
        train_pca_finale, test_pca_finale, numero_features, **migliori_iperparametri_quantistici
    )
    
    K_train_finale, K_test_finale = migliori_matrici_gram

    miglior_modello = SVC(kernel = "precomputed", **migliori_iperparametri_classici)
    miglior_modello.fit(K_train_finale, y_cv_raw)

    # Stampa nel terminale della miglior configurazione di iperparametri trovata, e relativo F1-score.

    if migliori_iperparametri_quantistici:
        for nome_iperparametro, valore_iperparametro in migliori_iperparametri_quantistici.items():
            print(f"Miglior iperparametro quantistico per {nome_iperparametro}: {valore_iperparametro}")
            
    if migliori_iperparametri_classici:
        for nome_iperparametro, valore_iperparametro in migliori_iperparametri_classici.items():
            print(f"Miglior iperparametro classico per {nome_iperparametro}: {valore_iperparametro}")
            
    print(f"F1-Score (Macro) del miglior modello in {numero_fold_effettivo}-fold CV: {miglior_score * 100:.2f}% (± {miglior_std_score * 100:.2f}%)\n")

    # Salvataggio dei file con i risultati di ogni combinazione di iperparametri (file .xlsx).
    
    df = pd.DataFrame(storico_risultati_addestramento)

    df["iperparametri_quantistici"] = df["iperparametri_quantistici"].astype(str)
    df["iperparametri_classici"] = df["iperparametri_classici"].astype(str)

    nome_file_excel = os.path.join(nome_cartella, "storico_risultati_"+nome_dataset+".xlsx")

    with pd.ExcelWriter(nome_file_excel, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name='Risultati')
        foglio = writer.sheets['Risultati']

        from openpyxl.utils import get_column_letter 
        from openpyxl.styles import PatternFill

        for indice_colonna, nome_colonna in enumerate(df.columns):

            lunghezza_massima = max(
                df[nome_colonna].astype(str).map(len).max(),
                len(nome_colonna)
            ) + 2

            lettera_colonna = get_column_letter(indice_colonna + 1)
            foglio.column_dimensions[lettera_colonna].width = lunghezza_massima

        indice_migliore = df["F1_score_macro(%)"].idxmax()
        riga_excel = indice_migliore + 2 # +1 per l'intestazione, +1 perché Excel conta partendo da 1
        riempimento_giallo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in range(1, len(df.columns) + 1):

            foglio.cell(row=riga_excel, column=col).fill = riempimento_giallo

    return miglior_set_adattato, migliori_matrici_gram, miglior_modello, storico_risultati_addestramento